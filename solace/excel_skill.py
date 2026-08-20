"""Offline Excel knowledge and safe workbook helpers for Solace."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import FORMULAE, range_boundaries


@dataclass(frozen=True)
class FormulaGuide:
    name: str
    syntax: str
    purpose: str
    example: str
    notes: str = ""


# openpyxl exposes a broad classic Excel function-name registry. These additions
# cover newer dynamic-array, lambda/helper, regex, and formula-aggregation names
# that may not exist in that registry on every openpyxl release.
MODERN_FUNCTIONS: Set[str] = {
    "BYCOL",
    "BYROW",
    "CHOOSECOLS",
    "CHOOSEROWS",
    "DROP",
    "EXPAND",
    "FILTER",
    "GROUPBY",
    "HSTACK",
    "IMAGE",
    "ISOMITTED",
    "LAMBDA",
    "LET",
    "MAKEARRAY",
    "MAP",
    "PIVOTBY",
    "RANDARRAY",
    "REDUCE",
    "REGEXEXTRACT",
    "REGEXREPLACE",
    "REGEXTEST",
    "SCAN",
    "SEQUENCE",
    "SORT",
    "SORTBY",
    "TAKE",
    "TEXTAFTER",
    "TEXTBEFORE",
    "TEXTSPLIT",
    "TOCOL",
    "TOROW",
    "TRIMRANGE",
    "UNIQUE",
    "VSTACK",
    "WRAPCOLS",
    "WRAPROWS",
    "XLOOKUP",
    "XMATCH",
}


def _guide(name: str, syntax: str, purpose: str, example: str, notes: str = "") -> FormulaGuide:
    return FormulaGuide(name, syntax, purpose, example, notes)


FORMULA_GUIDES: Dict[str, FormulaGuide] = {
    guide.name: guide
    for guide in [
        _guide("SUM", "SUM(number1, [number2], ...)", "Adds numbers or ranges.", "=SUM(B2:B20)"),
        _guide("AVERAGE", "AVERAGE(number1, [number2], ...)", "Returns the arithmetic mean.", "=AVERAGE(C2:C20)"),
        _guide("MIN", "MIN(number1, [number2], ...)", "Returns the smallest value.", "=MIN(D2:D100)"),
        _guide("MAX", "MAX(number1, [number2], ...)", "Returns the largest value.", "=MAX(D2:D100)"),
        _guide("COUNT", "COUNT(value1, [value2], ...)", "Counts numeric cells.", "=COUNT(B2:B100)"),
        _guide("COUNTA", "COUNTA(value1, [value2], ...)", "Counts nonblank cells.", "=COUNTA(A2:A100)"),
        _guide("COUNTBLANK", "COUNTBLANK(range)", "Counts blank cells.", "=COUNTBLANK(C2:C100)"),
        _guide(
            "IF",
            "IF(logical_test, value_if_true, value_if_false)",
            "Returns one value when a test is true and another when false.",
            '=IF(C2>=50,"Pass","Fail")',
        ),
        _guide(
            "IFS",
            "IFS(test1, value1, [test2, value2], ...)",
            "Evaluates multiple conditions in order.",
            '=IFS(B2>=80,"A",B2>=70,"B",B2>=60,"C",TRUE,"D")',
        ),
        _guide("AND", "AND(logical1, [logical2], ...)", "TRUE only when every condition is true.", '=AND(B2>=50,C2="Paid")'),
        _guide("OR", "OR(logical1, [logical2], ...)", "TRUE when any condition is true.", '=OR(D2="Lagos",D2="Abuja")'),
        _guide("NOT", "NOT(logical)", "Reverses TRUE/FALSE.", '=NOT(B2="Closed")'),
        _guide("IFERROR", "IFERROR(value, value_if_error)", "Returns a fallback when a formula errors.", "=IFERROR(A2/B2,0)"),
        _guide("IFNA", "IFNA(value, value_if_na)", "Handles only #N/A errors.", '=IFNA(XLOOKUP(E2,A:A,B:B),"Not found")'),
        _guide("SUMIF", "SUMIF(range, criteria, [sum_range])", "Adds values matching one condition.", '=SUMIF(A2:A100,"Lagos",B2:B100)'),
        _guide(
            "SUMIFS",
            "SUMIFS(sum_range, criteria_range1, criteria1, ...)",
            "Adds values matching multiple conditions.",
            '=SUMIFS(D:D,A:A,"Lagos",B:B,">=2026-01-01")',
        ),
        _guide("COUNTIF", "COUNTIF(range, criteria)", "Counts cells matching one condition.", '=COUNTIF(C2:C100,"Paid")'),
        _guide(
            "COUNTIFS",
            "COUNTIFS(criteria_range1, criteria1, ...)",
            "Counts rows matching multiple conditions.",
            '=COUNTIFS(A:A,"Lagos",C:C,"Paid")',
        ),
        _guide("AVERAGEIF", "AVERAGEIF(range, criteria, [average_range])", "Averages values matching one condition.", '=AVERAGEIF(A:A,"North",B:B)'),
        _guide(
            "AVERAGEIFS",
            "AVERAGEIFS(average_range, criteria_range1, criteria1, ...)",
            "Averages values matching multiple conditions.",
            '=AVERAGEIFS(D:D,A:A,"North",C:C,">0")',
        ),
        _guide(
            "XLOOKUP",
            "XLOOKUP(lookup_value, lookup_array, return_array, [if_not_found], [match_mode], [search_mode])",
            "Modern exact/flexible lookup that can search left or right.",
            '=XLOOKUP(E2,A2:A100,B2:B100,"Not found")',
        ),
        _guide(
            "VLOOKUP",
            "VLOOKUP(lookup_value, table_array, col_index_num, [range_lookup])",
            "Looks down the first column of a table and returns from a later column.",
            "=VLOOKUP(E2,A2:C100,3,FALSE)",
            "Prefer XLOOKUP when available.",
        ),
        _guide(
            "HLOOKUP",
            "HLOOKUP(lookup_value, table_array, row_index_num, [range_lookup])",
            "Looks across the first row of a table.",
            "=HLOOKUP(B1,A1:M5,4,FALSE)",
        ),
        _guide("INDEX", "INDEX(array, row_num, [column_num])", "Returns a value at a row/column position.", "=INDEX(B2:B100,MATCH(E2,A2:A100,0))"),
        _guide("MATCH", "MATCH(lookup_value, lookup_array, [match_type])", "Returns the position of a lookup value.", "=MATCH(E2,A2:A100,0)"),
        _guide(
            "XMATCH",
            "XMATCH(lookup_value, lookup_array, [match_mode], [search_mode])",
            "Modern MATCH with flexible matching/search direction.",
            "=XMATCH(E2,A2:A100,0)",
        ),
        _guide("FILTER", "FILTER(array, include, [if_empty])", "Returns only rows/columns meeting a condition.", '=FILTER(A2:D100,D2:D100="Open","None")'),
        _guide("UNIQUE", "UNIQUE(array, [by_col], [exactly_once])", "Returns distinct values or rows.", "=UNIQUE(A2:A100)"),
        _guide("SORT", "SORT(array, [sort_index], [sort_order], [by_col])", "Returns a sorted array.", "=SORT(A2:D100,2,-1)"),
        _guide(
            "SORTBY",
            "SORTBY(array, by_array1, [sort_order1], ...)",
            "Sorts an array using one or more external sort ranges.",
            "=SORTBY(A2:D100,D2:D100,-1)",
        ),
        _guide("SEQUENCE", "SEQUENCE(rows, [columns], [start], [step])", "Generates a numeric sequence as a spilled array.", "=SEQUENCE(12,1,1,1)"),
        _guide("LET", "LET(name1, value1, calculation_or_name2, ...)", "Names intermediate calculations inside one formula.", "=LET(total,SUM(B2:B20),total*0.075)"),
        _guide("LAMBDA", "LAMBDA([parameter1, ...], calculation)", "Creates reusable custom functions without VBA.", "=LAMBDA(x,x*1.075)(B2)"),
        _guide("TEXTJOIN", "TEXTJOIN(delimiter, ignore_empty, text1, ...)", "Joins many text values with a delimiter.", '=TEXTJOIN(", ",TRUE,A2:A10)'),
        _guide("CONCAT", "CONCAT(text1, [text2], ...)", "Combines text values.", '=CONCAT(A2," - ",B2)'),
        _guide("TEXTSPLIT", "TEXTSPLIT(text, col_delimiter, [row_delimiter], ...)", "Splits text into spilled columns/rows.", '=TEXTSPLIT(A2,",")'),
        _guide("TEXTBEFORE", "TEXTBEFORE(text, delimiter, [instance_num], ...)", "Returns text before a delimiter.", '=TEXTBEFORE(A2,"@")'),
        _guide("TEXTAFTER", "TEXTAFTER(text, delimiter, [instance_num], ...)", "Returns text after a delimiter.", '=TEXTAFTER(A2,"@")'),
        _guide("LEFT", "LEFT(text, [num_chars])", "Returns characters from the left.", "=LEFT(A2,3)"),
        _guide("RIGHT", "RIGHT(text, [num_chars])", "Returns characters from the right.", "=RIGHT(A2,4)"),
        _guide("MID", "MID(text, start_num, num_chars)", "Returns characters from the middle.", "=MID(A2,3,5)"),
        _guide("LEN", "LEN(text)", "Counts characters.", "=LEN(A2)"),
        _guide("TRIM", "TRIM(text)", "Removes extra ASCII spaces.", "=TRIM(A2)"),
        _guide("CLEAN", "CLEAN(text)", "Removes nonprinting characters.", "=CLEAN(A2)"),
        _guide("SUBSTITUTE", "SUBSTITUTE(text, old_text, new_text, [instance_num])", "Replaces matching text.", '=SUBSTITUTE(A2,"-","")'),
        _guide("REPLACE", "REPLACE(old_text, start_num, num_chars, new_text)", "Replaces characters by position.", '=REPLACE(A2,1,3,"NEW")'),
        _guide("FIND", "FIND(find_text, within_text, [start_num])", "Case-sensitive text position search.", '=FIND("@",A2)'),
        _guide("SEARCH", "SEARCH(find_text, within_text, [start_num])", "Case-insensitive text position search.", '=SEARCH("lagos",A2)'),
        _guide("TEXT", "TEXT(value, format_text)", "Formats a number/date as text.", '=TEXT(A2,"dd-mmm-yyyy")'),
        _guide("VALUE", "VALUE(text)", "Converts numeric text to a number.", "=VALUE(A2)"),
        _guide("DATE", "DATE(year, month, day)", "Builds a date from components.", "=DATE(2026,8,20)"),
        _guide("TODAY", "TODAY()", "Returns the current date.", "=TODAY()"),
        _guide("NOW", "NOW()", "Returns the current date and time.", "=NOW()"),
        _guide("DATEDIF", "DATEDIF(start_date, end_date, unit)", "Returns elapsed years/months/days between dates.", '=DATEDIF(A2,B2,"Y")'),
        _guide("EOMONTH", "EOMONTH(start_date, months)", "Returns the final day of a month offset.", "=EOMONTH(A2,0)"),
        _guide("EDATE", "EDATE(start_date, months)", "Moves a date by whole months.", "=EDATE(A2,3)"),
        _guide(
            "NETWORKDAYS",
            "NETWORKDAYS(start_date, end_date, [holidays])",
            "Counts working days excluding weekends and optional holidays.",
            "=NETWORKDAYS(A2,B2,H2:H20)",
        ),
        _guide("WORKDAY", "WORKDAY(start_date, days, [holidays])", "Returns a date a number of working days away.", "=WORKDAY(A2,10,H2:H20)"),
        _guide("ROUND", "ROUND(number, num_digits)", "Rounds to a specified number of digits.", "=ROUND(B2,2)"),
        _guide("ROUNDUP", "ROUNDUP(number, num_digits)", "Rounds away from zero.", "=ROUNDUP(B2,0)"),
        _guide("ROUNDDOWN", "ROUNDDOWN(number, num_digits)", "Rounds toward zero.", "=ROUNDDOWN(B2,0)"),
        _guide("ABS", "ABS(number)", "Returns the absolute value.", "=ABS(B2)"),
        _guide("MOD", "MOD(number, divisor)", "Returns the remainder after division.", "=MOD(A2,2)"),
        _guide(
            "SUMPRODUCT",
            "SUMPRODUCT(array1, [array2], ...)",
            "Multiplies corresponding values and sums the products; useful for weighted calculations.",
            "=SUMPRODUCT(B2:B10,C2:C10)",
        ),
        _guide(
            "SUBTOTAL",
            "SUBTOTAL(function_num, ref1, [ref2], ...)",
            "Aggregates while optionally ignoring filtered/hidden rows.",
            "=SUBTOTAL(9,B2:B100)",
        ),
        _guide(
            "AGGREGATE",
            "AGGREGATE(function_num, options, array, [k])",
            "Performs robust aggregation while ignoring selected errors/hidden rows.",
            "=AGGREGATE(9,5,B2:B100)",
        ),
        _guide("NPV", "NPV(rate, value1, [value2], ...)", "Calculates net present value of periodic cash flows.", "=NPV(B1,C2:C10)+C1"),
        _guide("IRR", "IRR(values, [guess])", "Returns the internal rate of return for periodic cash flows.", "=IRR(B2:B12)"),
        _guide("PMT", "PMT(rate, nper, pv, [fv], [type])", "Calculates a periodic loan/investment payment.", "=PMT(12%/12,36,-500000)"),
        _guide("RANK.EQ", "RANK.EQ(number, ref, [order])", "Ranks a number within a list.", "=RANK.EQ(B2,$B$2:$B$100,0)"),
        _guide("PERCENTILE.INC", "PERCENTILE.INC(array, k)", "Returns an inclusive percentile.", "=PERCENTILE.INC(B2:B100,0.9)"),
        _guide("STDEV.S", "STDEV.S(number1, [number2], ...)", "Estimates sample standard deviation.", "=STDEV.S(B2:B100)"),
        _guide("CORREL", "CORREL(array1, array2)", "Returns the correlation coefficient.", "=CORREL(B2:B100,C2:C100)"),
        _guide("REGEXTEST", "REGEXTEST(text, pattern, [case_sensitivity])", "Tests whether text matches a regular expression.", '=REGEXTEST(A2,"^[0-9]{11}$")'),
        _guide(
            "REGEXEXTRACT",
            "REGEXEXTRACT(text, pattern, [return_mode], [case_sensitivity])",
            "Extracts regex matches from text.",
            '=REGEXEXTRACT(A2,"[0-9]+")',
        ),
        _guide(
            "REGEXREPLACE",
            "REGEXREPLACE(text, pattern, replacement, [occurrence], [case_sensitivity])",
            "Replaces regex matches.",
            '=REGEXREPLACE(A2,"\\s+"," ")',
        ),
        _guide("VSTACK", "VSTACK(array1, [array2], ...)", "Stacks arrays vertically.", "=VSTACK(A2:C10,E2:G10)"),
        _guide("HSTACK", "HSTACK(array1, [array2], ...)", "Stacks arrays horizontally.", "=HSTACK(A2:A10,C2:C10)"),
        _guide("TAKE", "TAKE(array, rows, [columns])", "Returns leading/trailing rows or columns.", "=TAKE(A2:D100,10)"),
        _guide("DROP", "DROP(array, rows, [columns])", "Removes leading/trailing rows or columns.", "=DROP(A2:D100,1)"),
        _guide(
            "GROUPBY",
            "GROUPBY(row_fields, values, function, [field_headers], [total_depth], [sort_order], [filter_array], [field_relationship])",
            "Creates a formula-driven grouped summary.",
            "=GROUPBY(A2:A100,D2:D100,SUM)",
        ),
        _guide(
            "PIVOTBY",
            "PIVOTBY(row_fields, col_fields, values, function, ...)",
            "Creates a formula-driven two-dimensional summary.",
            "=PIVOTBY(A2:A100,B2:B100,D2:D100,SUM)",
        ),
    ]
}


PIVOT_GUIDE = """[bold]PivotTable setup[/bold]
1. Put the source data in a clean rectangle with one header row and no merged headers.
2. Click any source cell, then Insert → PivotTable.
3. Confirm the table/range and choose a new or existing worksheet.
4. Drag categorical fields to Rows/Columns, numeric fields to Values, and optional fields to Filters.
5. Use Value Field Settings to choose Sum, Count, Average, Min, Max, and number formatting.
6. Group dates by month/quarter/year when useful, then Refresh after source data changes.

Tip: converting the source to an Excel Table first makes the source range expand more reliably.
Solace can create a pivot-style summary sheet with `/excel summarize ...`, but openpyxl does not create new native Excel PivotTables; it preserves existing ones."""

CHART_GUIDE = """[bold]Excel chart setup[/bold]
1. Select the clean data range, including headers.
2. Insert → Recommended Charts, or choose the chart type directly.
3. Use column/bar for category comparisons, line for ordered time trends, pie only for a small part-to-whole view, and scatter for two numeric variables.
4. Add a descriptive title and axis labels; remove decorative clutter.
5. If the source is an Excel Table, charts usually expand more cleanly as rows are added.

Solace can create bar, line, and pie charts in `.xlsx` workbooks with `/excel make-chart ...`."""

TABLE_GUIDE = """[bold]Excel Table setup[/bold]
Select the data → Ctrl+T (or Insert → Table) → confirm headers. Tables provide filters, structured references, consistent formatting, and ranges that grow as rows are added."""

CONDITIONAL_FORMAT_GUIDE = """[bold]Conditional formatting[/bold]
Select the target cells → Home → Conditional Formatting. Use highlight rules for thresholds/duplicates/dates, data bars for magnitude, color scales for relative intensity, or New Rule for formula-based logic. Keep rules few and meaningful."""

DATA_VALIDATION_GUIDE = """[bold]Data validation[/bold]
Select cells → Data → Data Validation. Choose Whole number, Decimal, Date, Time, Text length, List, or Custom. For controlled categories, use a List backed by a range/table rather than typing values repeatedly."""


EXCEL_ERROR_GUIDES: Dict[str, str] = {
    "#N/A": "Usually a lookup could not find a match. Check spaces/types and exact-match settings; use IFNA only after fixing the lookup logic.",
    "#VALUE!": "A formula received the wrong value type. Look for text where numbers/dates are expected or mismatched array sizes.",
    "#REF!": "A reference points to cells/ranges that were deleted or became invalid.",
    "#DIV/0!": "The denominator is zero or blank. Fix the data or guard intentionally with IF/IFERROR.",
    "#NAME?": "Excel does not recognize a function/name. Check spelling, quotes, version support, and named ranges.",
    "#NUM!": "A numeric calculation is invalid or outside the supported range.",
    "#SPILL!": "A dynamic-array result cannot expand because cells/merged ranges/tables block the spill area.",
    "#CALC!": "A dynamic-array/LAMBDA calculation produced an unsupported or empty-array state.",
}


def known_formula_names() -> List[str]:
    """Return the local Excel function-name index."""

    return sorted({str(name).upper() for name in FORMULAE} | MODERN_FUNCTIONS | set(FORMULA_GUIDES))


def search_functions(query: str, limit: int = 30) -> List[str]:
    """Search the formula-name index by name fragment."""

    term = query.strip().upper()
    names = known_formula_names()
    if not term:
        return names[:limit]
    exact = [name for name in names if name == term]
    starts = [name for name in names if name.startswith(term) and name not in exact]
    contains = [name for name in names if term in name and name not in exact and name not in starts]
    return (exact + starts + contains)[:limit]


def formula_guide(name: str) -> Optional[FormulaGuide]:
    return FORMULA_GUIDES.get(name.strip().upper())


def answer_excel_query(query: str) -> Optional[str]:
    """Return a deterministic offline answer when the query matches known Excel topics."""

    text = query.strip()
    lowered = text.lower()
    upper = text.upper()

    for error, explanation in EXCEL_ERROR_GUIDES.items():
        if error in upper:
            return "[bold]{}[/bold]\n{}".format(error, explanation)

    if "pivot" in lowered:
        return PIVOT_GUIDE
    if "conditional format" in lowered:
        return CONDITIONAL_FORMAT_GUIDE
    if "data validation" in lowered or "dropdown" in lowered:
        return DATA_VALIDATION_GUIDE
    if "excel table" in lowered or lowered in {"table", "tables"}:
        return TABLE_GUIDE
    if "chart" in lowered or "graph" in lowered:
        return CHART_GUIDE

    tokens = [token.strip("()=,.;:[]{}\"'?!").upper() for token in text.split()]
    for token in tokens:
        guide = FORMULA_GUIDES.get(token)
        if guide:
            parts = [
                "[bold]{}[/bold]".format(guide.name),
                guide.purpose,
                "Syntax: [cyan]{}[/cyan]".format(guide.syntax),
                "Example: [green]{}[/green]".format(guide.example),
            ]
            if guide.notes:
                parts.append("Note: {}".format(guide.notes))
            return "\n".join(parts)
    return None


def qwen_excel_prompt(query: str) -> str:
    """Build a constrained local-model prompt for Excel questions outside deterministic guides."""

    return (
        "You are Solace's Excel specialist. Answer the user's Excel question accurately and practically. "
        "Use modern Excel syntax when appropriate, state version caveats, and give formulas in English "
        "function names with comma separators. Never invent a function. If a native PivotTable or feature "
        "must be created in Excel's UI, say so clearly. Give concise step-by-step instructions.\n\n"
        "User question: " + query.strip()
    )


def _workbook_output(source: Path, suffix: str, output: Optional[Path]) -> Path:
    if output is not None:
        target = output.expanduser()
    else:
        target = source.with_name("{}-{}{}".format(source.stem, suffix, source.suffix))
    if target.exists():
        raise FileExistsError("Output already exists: {}".format(target))
    target.parent.mkdir(parents=True, exist_ok=True)
    return target


def _load(path: Path):
    source = path.expanduser()
    if not source.is_file():
        raise FileNotFoundError(source)
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("Solace currently edits .xlsx/.xlsm workbooks, not {}".format(source.suffix or "this file"))
    return source, load_workbook(source, keep_vba=source.suffix.lower() == ".xlsm")


def inspect_workbook(path: Path) -> List[Dict[str, object]]:
    """Inspect sheets, dimensions, formulas, tables, charts, and existing pivots."""

    source, workbook = _load(path)
    rows: List[Dict[str, object]] = []
    try:
        for sheet in workbook.worksheets:
            formula_count = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
            rows.append(
                {
                    "sheet": sheet.title,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                    "formulas": formula_count,
                    "tables": len(sheet.tables),
                    "charts": len(getattr(sheet, "_charts", [])),
                    "pivots": len(getattr(sheet, "_pivots", [])),
                }
            )
    finally:
        workbook.close()
    if not rows:
        raise ValueError("Workbook has no worksheets: {}".format(source))
    return rows


def set_formula(path: Path, sheet_name: str, cell: str, formula: str, output: Optional[Path] = None) -> Path:
    """Write one formula to a new workbook copy; never overwrite the source."""

    source, workbook = _load(path)
    target = _workbook_output(source, "formula", output)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError("Worksheet not found: {}".format(sheet_name))
        value = formula.strip()
        if not value.startswith("="):
            value = "=" + value
        workbook[sheet_name][cell.upper()] = value
        workbook.save(target)
    finally:
        workbook.close()
    return target


def create_chart(
    path: Path,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    output: Optional[Path] = None,
    anchor: str = "H2",
    title: Optional[str] = None,
) -> Path:
    """Create a simple bar/line/pie chart in a new workbook copy."""

    source, workbook = _load(path)
    target = _workbook_output(source, "chart", output)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError("Worksheet not found: {}".format(sheet_name))
        sheet = workbook[sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(data_range.upper())
        if max_row <= min_row:
            raise ValueError("Chart range needs a header row plus at least one data row.")
        if max_col <= min_col:
            raise ValueError("Chart range needs a category column plus at least one value column.")

        kind = chart_type.strip().lower()
        if kind in {"bar", "column"}:
            chart = BarChart()
            if kind == "bar":
                chart.type = "bar"
        elif kind == "line":
            chart = LineChart()
        elif kind == "pie":
            chart = PieChart()
        else:
            raise ValueError("Supported chart types: bar, column, line, pie")

        categories = Reference(sheet, min_col=min_col, min_row=min_row + 1, max_row=max_row)
        if kind == "pie":
            data = Reference(sheet, min_col=min_col + 1, min_row=min_row, max_row=max_row)
        else:
            data = Reference(sheet, min_col=min_col + 1, max_col=max_col, min_row=min_row, max_row=max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.title = title or "Solace chart"
        chart.height = 8
        chart.width = 14
        sheet.add_chart(chart, anchor.upper())
        workbook.save(target)
    finally:
        workbook.close()
    return target


def create_summary(
    path: Path,
    sheet_name: str,
    group_header: str,
    value_header: str,
    operation: str = "sum",
    output: Optional[Path] = None,
) -> Path:
    """Create a pivot-style grouped summary sheet without pretending it is a native PivotTable."""

    source, workbook = _load(path)
    target = _workbook_output(source, "summary", output)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError("Worksheet not found: {}".format(sheet_name))
        source_sheet = workbook[sheet_name]
        headers = {
            str(cell.value).strip().lower(): cell.column
            for cell in source_sheet[1]
            if cell.value is not None and str(cell.value).strip()
        }
        group_key = group_header.strip().lower()
        value_key = value_header.strip().lower()
        if group_key not in headers:
            raise KeyError("Column header not found: {}".format(group_header))
        if value_key not in headers:
            raise KeyError("Column header not found: {}".format(value_header))

        op = operation.strip().lower()
        if op not in {"sum", "count", "average", "avg"}:
            raise ValueError("Summary operation must be sum, count, or average.")

        grouped: Dict[str, Tuple[float, int]] = {}
        group_col = headers[group_key]
        value_col = headers[value_key]
        for row in range(2, source_sheet.max_row + 1):
            group_value = source_sheet.cell(row=row, column=group_col).value
            value = source_sheet.cell(row=row, column=value_col).value
            if group_value is None:
                continue
            key = str(group_value)
            total, count = grouped.get(key, (0.0, 0))
            if op == "count":
                grouped[key] = (total, count + (0 if value is None else 1))
                continue
            if not isinstance(value, (int, float)):
                continue
            grouped[key] = (total + float(value), count + 1)

        title = "Solace Summary"
        if title in workbook.sheetnames:
            del workbook[title]
        summary = workbook.create_sheet(title)
        result_label = "Average" if op in {"average", "avg"} else op.title()
        summary.append([group_header, "{} of {}".format(result_label, value_header)])
        for key in sorted(grouped, key=lambda item: item.casefold()):
            total, count = grouped[key]
            result: float
            if op == "count":
                result = float(count)
            elif op in {"average", "avg"}:
                result = total / count if count else 0.0
            else:
                result = total
            summary.append([key, result])
        summary.freeze_panes = "A2"
        workbook.save(target)
    finally:
        workbook.close()
    return target


__all__ = [
    "CHART_GUIDE",
    "FORMULA_GUIDES",
    "PIVOT_GUIDE",
    "FormulaGuide",
    "answer_excel_query",
    "create_chart",
    "create_summary",
    "formula_guide",
    "inspect_workbook",
    "known_formula_names",
    "qwen_excel_prompt",
    "search_functions",
    "set_formula",
]
