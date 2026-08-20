from pathlib import Path

from openpyxl import Workbook, load_workbook

from solace.excel_skill import (
    answer_excel_query,
    create_chart,
    create_summary,
    inspect_workbook,
    known_formula_names,
    search_functions,
    set_formula,
)


def _sample_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sales"
    sheet.append(["Region", "Revenue", "Units"])
    sheet.append(["Lagos", 100, 2])
    sheet.append(["Abuja", 80, 4])
    sheet.append(["Lagos", 50, 1])
    sheet["D1"] = "Double"
    sheet["D2"] = "=B2*2"
    workbook.save(path)
    workbook.close()


def test_formula_index_includes_classic_and_modern_functions() -> None:
    names = known_formula_names()
    assert "SUM" in names
    assert "XLOOKUP" in names
    assert "FILTER" in names
    assert search_functions("look")[:2] == ["LOOKUP", "XLOOKUP"] or "XLOOKUP" in search_functions("look")


def test_excel_query_returns_deterministic_formula_and_pivot_guides() -> None:
    xlookup = answer_excel_query("How do I use XLOOKUP?")
    pivot = answer_excel_query("help me build a pivot table")

    assert xlookup is not None
    assert "XLOOKUP" in xlookup
    assert "lookup_value" in xlookup
    assert pivot is not None
    assert "PivotTable setup" in pivot
    assert "does not create new native Excel PivotTables" in pivot
    assert "/excel summarize" in pivot


def test_inspect_and_set_formula_create_safe_copy(tmp_path: Path) -> None:
    source = tmp_path / "sales.xlsx"
    _sample_workbook(source)

    inspection = inspect_workbook(source)
    assert inspection[0]["sheet"] == "Sales"
    assert inspection[0]["formulas"] == 1

    output = set_formula(source, "Sales", "E2", "SUM(B2:C2)")
    assert output.exists()
    assert output != source

    workbook = load_workbook(output, data_only=False)
    try:
        assert workbook["Sales"]["E2"].value == "=SUM(B2:C2)"
    finally:
        workbook.close()


def test_create_chart_and_pivot_style_summary(tmp_path: Path) -> None:
    source = tmp_path / "sales.xlsx"
    _sample_workbook(source)

    chart_output = create_chart(source, "Sales", "A1:B4", "bar")
    assert chart_output.exists()
    chart_book = load_workbook(chart_output)
    try:
        assert len(chart_book["Sales"]._charts) == 1
    finally:
        chart_book.close()

    summary_output = create_summary(source, "Sales", "Region", "Revenue", "sum")
    assert summary_output.exists()
    summary_book = load_workbook(summary_output, data_only=False)
    try:
        summary = summary_book["Solace Summary"]
        values = {summary.cell(row=row, column=1).value: summary.cell(row=row, column=2).value for row in range(2, 4)}
        assert values == {"Abuja": 80, "Lagos": 150}
    finally:
        summary_book.close()
